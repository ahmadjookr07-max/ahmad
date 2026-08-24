# -*- coding: utf-8 -*-
"""delivery_excel_report — تقرير Excel اختياري داخل حزمة الصور الجاهزة.

يبني قائمة قابلة للفرز للصور التي تم ربطها فعلًا بملف الأصناف. لا يضع
صفوف المراجعة أو الصور غير المرتبطة في تقرير التسليم، ولا يخمّن باركودًا
عند تعدد المرشحين في Excel.
"""
from __future__ import annotations

import os
import tempfile
from datetime import datetime
from pathlib import Path
from typing import Any

REPORT_FILENAME = "قائمة_الصور_المرتبطة.xlsx"
LINKED_STATUSES = {"matched", "manual"}

__all__ = [
    "REPORT_FILENAME",
    "linked_delivery_rows",
    "write_delivery_excel_report",
]


def _resolve(raw: Any, workspace: Path | None) -> Path | None:
    text = str(raw or "").strip()
    if not text:
        return None
    path = Path(text)
    if not path.is_absolute() and workspace is not None:
        candidate = workspace / path
        if candidate.is_file():
            return candidate
    return path if path.is_file() else None


def _normalise_unit(value: object) -> str:
    try:
        from engine_v2.catalog_index_v2 import normalize_text
        return normalize_text(str(value or "")).replace(" ", "")
    except Exception:
        return str(value or "").strip().replace(" ", "").casefold()


def _output_unit(path: Path) -> str:
    """يستخرج الوحدة كما تظهر في اسم الصورة النهائي، إن أمكن."""
    try:
        from engine_v2.naming_v2 import parse_name
        parsed = parse_name(path.stem)
        return str(getattr(parsed, "unit", "") or "").strip() if parsed else ""
    except Exception:
        return ""


def _catalog_data(index: Any, item_code: str, unit: str,
                  observed_barcode: str) -> dict[str, str]:
    """يُكمل الحقول من Excel بلا اختيار باركود ملتبس.

    الباركود الموجود في نتيجة الربط هو الأولوية. إن غاب، يؤخذ فقط من
    صف Excel وحيد متوافق مع رقم الصنف والوحدة؛ تعدد المرشحين يبقي الحقل
    فارغًا بدل تقديم معلومة قد تخص عبوة مختلفة.
    """
    data = {"name": "", "unit": unit, "barcode": observed_barcode}
    if index is None or not item_code:
        return data
    try:
        rows = list(index.rows_for_code(item_code))
    except Exception:
        return data
    if not rows:
        return data
    unit_key = _normalise_unit(unit)
    scoped = [row for row in rows
              if unit_key and _normalise_unit(row.get("unit", "")) == unit_key]
    candidates = scoped or rows
    # الاسم والوحدة يظلان من صفٍّ متوافق إن وجد، وإلا من الصف الأول فقط.
    row = candidates[0]
    data["name"] = str(row.get("name", "") or "").strip()
    data["unit"] = str(row.get("unit", "") or "").strip() or unit
    if not data["barcode"]:
        barcodes = {
            str(candidate.get("barcode", "") or "").strip()
            for candidate in candidates
            if str(candidate.get("barcode", "") or "").strip()
        }
        if len(barcodes) == 1:
            data["barcode"] = next(iter(barcodes))
    return data


def linked_delivery_rows(result: Any, workspace: Path | None = None,
                         catalog_index: Any = None) -> list[dict[str, str]]:
    """يبني صفوف التقرير من المنتجات المرتبطة ذات مخرجات موجودة فقط."""
    rows: list[dict[str, str]] = []
    seen_outputs: set[str] = set()
    for item in getattr(result, "items", []) or []:
        status = str(getattr(item, "status", "") or "").strip().casefold()
        if status not in LINKED_STATUSES:
            continue
        output = _resolve(getattr(item, "output_path", ""), workspace)
        if output is None:
            continue
        output_key = os.path.normcase(os.path.normpath(str(output)))
        if output_key in seen_outputs:
            continue
        seen_outputs.add(output_key)
        item_code = str(getattr(item, "item_code", "") or "").strip()
        barcode = str(getattr(item, "barcode", "") or "").strip()
        unit = _output_unit(output)
        catalog = _catalog_data(catalog_index, item_code, unit, barcode)
        product_name = str(getattr(item, "product_name", "") or "").strip()
        source_name = str(getattr(item, "source_name", "") or "").strip()
        match_source = str(getattr(item, "match_source", "") or "").strip()
        rows.append({
            "item_code": item_code,
            "product_name": product_name or catalog["name"],
            "barcode": barcode or catalog["barcode"],
            "unit": catalog["unit"] or unit,
            "output_name": output.name,
            "archive_path": f"processed/{output.name}",
            "status": "ربط يدوي" if status == "manual" else "ربط تلقائي",
            "match_source": match_source,
            "source_name": source_name,
        })
    return sorted(rows, key=lambda row: (row["item_code"], row["output_name"].casefold()))


def write_delivery_excel_report(result: Any, workspace: Path | None = None,
                                catalog_index: Any = None) -> Path | None:
    """يكتب تقرير Excel ذريًا داخل ``reports/`` ويرجع مساره عند النجاح."""
    if workspace is None:
        raw = str(getattr(result, "workspace", "") or "").strip()
        workspace = Path(raw) if raw else None
    if workspace is None:
        return None
    workspace = Path(workspace)
    rows = linked_delivery_rows(result, workspace, catalog_index)
    reports_dir = workspace / "reports"
    reports_dir.mkdir(parents=True, exist_ok=True)
    target = reports_dir / REPORT_FILENAME

    try:
        from openpyxl import Workbook
        from openpyxl.styles import Alignment, Border, Font, PatternFill, Side
        from openpyxl.utils import get_column_letter

        wb = Workbook()
        ws = wb.active
        ws.title = "الصور المرتبطة"
        ws.sheet_view.rightToLeft = True
        ws.sheet_view.showGridLines = False
        ws.column_dimensions["A"].width = 3

        primary = "1F4E79"
        light = "D6E3F0"
        ws.merge_cells("B2:J2")
        title = ws["B2"]
        title.value = "قائمة الصور الجاهزة المرتبطة بملف الأصناف"
        title.font = Font(name="Calibri", size=16, bold=True, color="FFFFFF")
        title.fill = PatternFill("solid", fgColor=primary)
        title.alignment = Alignment(horizontal="center", vertical="center")
        ws.row_dimensions[2].height = 30
        ws.merge_cells("B3:J3")
        subtitle = ws["B3"]
        subtitle.value = "تتضمن الصور التي رُبطت تلقائيًا أو يدويًا فقط؛ مصدر بيانات الصنف هو نتيجة الربط وملف Excel المختار."
        subtitle.font = Font(name="Calibri", size=10, color="44546A")
        subtitle.alignment = Alignment(horizontal="right", vertical="center", wrap_text=True)
        ws.row_dimensions[3].height = 28
        ws["B5"] = "عدد الصور الجاهزة المرتبطة"
        ws["B5"].font = Font(name="Calibri", bold=True, color="FFFFFF")
        ws["B5"].fill = PatternFill("solid", fgColor=primary)
        ws["C5"] = len(rows)
        ws["C5"].font = Font(name="Calibri", bold=True, color=primary)
        ws["C5"].fill = PatternFill("solid", fgColor="E8F5E9")
        ws["E5"] = "تاريخ إنشاء التقرير"
        ws["E5"].font = Font(name="Calibri", bold=True, color="FFFFFF")
        ws["E5"].fill = PatternFill("solid", fgColor=primary)
        ws["F5"] = datetime.now().strftime("%Y-%m-%d %H:%M")
        ws["F5"].font = Font(name="Calibri", color="44546A")

        headers = [
            "رقم الصنف", "اسم الصنف", "الباركود الخطي", "الوحدة",
            "اسم الصورة الجاهزة", "المسار داخل الحزمة", "حالة الربط",
            "مصدر المطابقة", "اسم الصورة الأصلية",
        ]
        header_row = 7
        for col, header in enumerate(headers, start=2):
            cell = ws.cell(header_row, col, header)
            cell.font = Font(name="Calibri", size=10, bold=True, color="FFFFFF")
            cell.fill = PatternFill("solid", fgColor=primary)
            cell.alignment = Alignment(horizontal="center", vertical="center", wrap_text=True)
        ws.row_dimensions[header_row].height = 32

        keys = [
            "item_code", "product_name", "barcode", "unit", "output_name",
            "archive_path", "status", "match_source", "source_name",
        ]
        thin = Side(style="thin", color="D9E2F3")
        for row_number, record in enumerate(rows, start=header_row + 1):
            tint = "F7FBFF" if row_number % 2 == 0 else "FFFFFF"
            for col, key in enumerate(keys, start=2):
                cell = ws.cell(row_number, col, record[key])
                cell.font = Font(name="Calibri", size=10, color="1F1F1F")
                cell.fill = PatternFill("solid", fgColor=tint)
                alignment = "left" if key in {"product_name", "output_name", "archive_path", "source_name"} else "center"
                cell.alignment = Alignment(horizontal=alignment, vertical="center", wrap_text=key == "product_name")
                cell.border = Border(top=thin, bottom=thin)
            ws.row_dimensions[row_number].height = 20

        last_row = max(header_row, header_row + len(rows))
        ws.auto_filter.ref = f"B{header_row}:J{last_row}"
        ws.freeze_panes = "B8"
        widths = [16, 34, 18, 13, 34, 38, 14, 18, 28]
        for offset, width in enumerate(widths, start=2):
            ws.column_dimensions[get_column_letter(offset)].width = width
        ws.sheet_properties.pageSetUpPr.fitToPage = True
        ws.page_setup.orientation = "landscape"
        ws.page_setup.fitToWidth = 1
        ws.page_margins.right = 0.25
        ws.page_margins.left = 0.25
        ws.page_margins.top = 0.45
        ws.page_margins.bottom = 0.45

        handle, temp_name = tempfile.mkstemp(prefix=f".{target.stem}.", suffix=".tmp", dir=str(reports_dir))
        os.close(handle)
        temporary = Path(temp_name)
        try:
            wb.save(temporary)
            os.replace(temporary, target)
        finally:
            temporary.unlink(missing_ok=True)
        try:
            setattr(result, "delivery_excel_report", str(target))
        except Exception:
            # نتائج بعض المحركات frozen dataclasses؛ تسجيل التقرير لا يجب
            # أن يفشل بسبب ذلك لأن حزمة ZIP تحتاج مساره بعد الإنشاء.
            try:
                object.__setattr__(result, "delivery_excel_report", str(target))
            except Exception:
                pass
        return target
    except Exception:
        return None
