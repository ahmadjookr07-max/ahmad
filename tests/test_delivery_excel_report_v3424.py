# -*- coding: utf-8 -*-
"""حارس v3.4.24: تقرير Excel للصور الجاهزة المرتبطة داخل حزمة ZIP."""
from __future__ import annotations

import os
import sys
import tempfile
import zipfile
from pathlib import Path
from types import SimpleNamespace

import cv2
import numpy as np
from openpyxl import load_workbook

ROOT = Path(__file__).resolve().parents[1]
sys.path[:0] = [str(ROOT / "windows_app"), str(ROOT / "src")]

from delivery_excel_report import REPORT_FILENAME, linked_delivery_rows, write_delivery_excel_report
from delivery_zip_fast import write_delivery_zip

FAILS: list[str] = []


def check(condition: bool, name: str) -> None:
    print(f"{'PASS' if condition else 'FAIL'}  {name}")
    if not condition:
        FAILS.append(name)


class Catalog:
    def rows_for_code(self, code: str):
        return {
            "100001": [{"code": "100001", "name": "قهوة مختصة محمصة", "unit": "حبه", "barcode": "6281111111111"}],
            "100002": [{"code": "100002", "name": "تمر فاخر", "unit": "كرتون", "barcode": "6282222222222"}],
        }.get(code, [])


with tempfile.TemporaryDirectory(prefix="delivery_excel_") as td:
    workspace = Path(td)
    processed = workspace / "processed"
    processed.mkdir()
    image = np.full((30, 40, 3), 230, np.uint8)
    front = processed / "100001_حبه.webp"
    back = processed / "100002_كرتون-2.webp"
    cv2.imwrite(str(front), image, [cv2.IMWRITE_WEBP_QUALITY, 101])
    cv2.imwrite(str(back), image, [cv2.IMWRITE_WEBP_QUALITY, 101])

    items = [
        SimpleNamespace(source_name="front.jpg", status="manual", item_code="100001",
                        product_name="", barcode="", output_path="processed/100001_حبه.webp",
                        match_source="manual_excel"),
        # المخرج المكرر يجب ألا يظهر مرتين في التقرير أو ZIP.
        SimpleNamespace(source_name="front-copy.jpg", status="manual", item_code="100001",
                        product_name="قهوة مختصة محمصة", barcode="", output_path="processed/100001_حبه.webp",
                        match_source="manual_excel"),
        SimpleNamespace(source_name="back.jpg", status="matched", item_code="100002",
                        product_name="تمر فاخر", barcode="6282222222222", output_path="processed/100002_كرتون-2.webp",
                        match_source="linear_barcode"),
        # المراجعة غير المرتبطة لا تدخل تقرير التسليم حتى لو كان لها ملف.
        SimpleNamespace(source_name="review.jpg", status="review", item_code="",
                        product_name="", barcode="", output_path="processed/100002_كرتون-2.webp",
                        match_source=""),
    ]
    result = SimpleNamespace(items=items, workspace=str(workspace),
                             delivery_zip=str(workspace / "delivery.zip"),
                             report_json="", report_csv="")

    rows = linked_delivery_rows(result, workspace, Catalog())
    check(len(rows) == 2, "التقرير يضم الصور المرتبطة الجاهزة فقط بلا تكرار")
    check(rows[0]["item_code"] == "100001" and rows[0]["product_name"] == "قهوة مختصة محمصة",
          "اسم الصنف يُستكمل من فهرس Excel")
    check(rows[0]["unit"] == "حبه" and rows[0]["barcode"] == "6281111111111",
          "الوحدة والباركود يأتيان من صف Excel المتوافق")
    check(rows[1]["output_name"] == "100002_كرتون-2.webp",
          "اسم الصورة النهائية محفوظ كما هو")

    report = write_delivery_excel_report(result, workspace, Catalog())
    check(report is not None and report.is_file(), "ملف Excel يُنشأ في مجلد reports")
    check(getattr(result, "delivery_excel_report", "") == str(report),
          "مسار التقرير مسجل مع نتيجة الحزمة")

    wb = load_workbook(report, data_only=True)
    ws = wb["الصور المرتبطة"]
    check(ws["C5"].value == 2, "ملخص Excel يعرض العدد الصحيح للصور المرتبطة")
    check(ws["B7"].value == "رقم الصنف" and ws["F8"].value == "100001_حبه.webp",
          "Excel يحتوي أعمدة البيانات واسم الملف النهائي")
    check(ws["D8"].value == "6281111111111" and ws["E9"].value == "كرتون",
          "Excel يحفظ الباركود والوحدة الصحيحة")
    wb.close()

    check(write_delivery_zip(result, workspace), "تُكتب حزمة ZIP مع تقرير Excel")
    with zipfile.ZipFile(workspace / "delivery.zip") as archive:
        names = set(archive.namelist())
        check("processed/100001_حبه.webp" in names and "processed/100002_كرتون-2.webp" in names,
              "الصور الجاهزة محفوظة داخل الحزمة")
        check(f"reports/{REPORT_FILENAME}" in names,
              "تقرير Excel يظهر داخل مجلد reports في الحزمة")
        check(archive.testzip() is None, "حزمة ZIP سليمة")

    # إيقاف الخيار يمسح المرجع فقط؛ لا يحذف التقرير المحلي ولا الصور.
    result.delivery_excel_report = ""
    result.delivery_zip = str(workspace / "delivery_without_excel.zip")
    check(write_delivery_zip(result, workspace), "تُكتب الحزمة عند إيقاف التقرير")
    with zipfile.ZipFile(result.delivery_zip) as archive:
        check(f"reports/{REPORT_FILENAME}" not in archive.namelist(),
              "إيقاف الخيار يمنع إرفاق Excel في الحزمة التالية")

print("ALL:", "PASS" if not FAILS else "FAIL")
sys.exit(0 if not FAILS else 1)
